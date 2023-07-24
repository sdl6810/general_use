import os, sys
from openpyxl import load_workbook
from selenium import webdriver
import pandas as pd

from selenium.webdriver import Firefox
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.firefox.firefox_profile import FirefoxProfile
from selenium.webdriver.common.by import By

#code should determine when it should stop cycling through the column
def importXlColumnToPythonList(filepath, worksheetName, columnLetter, character):
	wbk = load_workbook(filepath)
	sht = wbk[worksheetName]

	data = []
	for i in range(0,(sht.max_row-(sht.min_row+1))):
		if sht[columnLetter + str((sht.min_row+2)+i)].value is None:
			data.append(character)
		else:
			data.append(sht[columnLetter+str((sht.min_row+2)+i)].value)
	return data

# def importXlTable(filepath, worksheetName):
# 	#function should test where the end of the column is

#imports table from Wikipedia and puts it in Excel or in pandas
def importWikiTableFromWeb(url, rows, cols):
	return None


def main():
	storedData = importXlColumnToPythonList('/home/sdl5384/Desktop/Python_SRC/Astronomy/lunar_phases.xlsx','periApo', 'B', 0)

main()